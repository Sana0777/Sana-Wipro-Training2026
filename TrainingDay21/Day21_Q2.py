import pandas as pd
from openpyxl import load_workbook

df = pd.read_excel("sales_summary.xlsx", sheet_name="Sheet1")

df["Total"] = df["Quantity"] * df["Price"]

df.to_excel("sales_summary1.xlsx", index=False)

print("sales_summary.xlsx created successfully!")


# using openpyxl only
wb = load_workbook("sales_summary.xlsx")
ws = wb["Sheet1"]

ws.cell(row=1, column=4).value = "Total"

for row in range(2, ws.max_row + 1):
    quantity = ws.cell(row=row, column=2).value
    price = ws.cell(row=row, column=3).value
    ws.cell(row=row, column=4).value = quantity * price

wb.save("sales_summary2.xlsx")

print("sales_summary.xlsx created successfully!")
